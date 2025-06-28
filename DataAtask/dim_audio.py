import pandas as pd
import os
import json
from openpyxl import load_workbook

input_file = r"C:\Users\user\Downloads\raw_data.xlsx"

output_file = r"C:\Users\user\Desktop\Docs\final_data1.xlsx"

output_dir = os.path.dirname(output_file)
if not os.path.exists(output_dir):
    try:
        os.makedirs(output_dir)
        print(f"Created directory: {output_dir}")
    except Exception as e:
        print(f"Error creating directory {output_dir}: {e}")
        exit(1)

try:
    df = pd.read_excel(input_file, sheet_name='Sheet1')
except FileNotFoundError:
    print(f"Error: Input file '{input_file}' not found.")
    exit(1)
except Exception as e:
    print(f"Error reading input file '{input_file}': {e}")
    exit(1)

audio_data = []

seen_urls = set()
for index, row in df.iterrows():
    try:
        content = json.loads(row['raw_content'])
        if 'meeting_attendees' in content:
            seen_emails = set()
            unique_attendees = []
            for attendee in content['meeting_attendees']:
                email = attendee.get('email')
                if email and email not in seen_emails:
                    seen_emails.add(email)
                    unique_attendees.append(attendee)
            content['meeting_attendees'] = unique_attendees
        audio_url = content.get('audio_url', 'No URL')
        if audio_url not in seen_urls:
            seen_urls.add(audio_url)
            audio_data.append({
                'raw_audio_url': audio_url,
                'audio_id': len(seen_urls)
            })
    except json.JSONDecodeError as e:
        print(f"Warning: Could not parse JSON in row {index + 2} (source_id: {row['source_id']}): {row['raw_content'][:50]}... (Error: {e})")
        continue
    except Exception as e:
        print(f"Unexpected error in row {index + 2} (source_id: {row['source_id']}): {e}")
        continue

audio_df = pd.DataFrame(audio_data)

try:
    if os.path.exists(output_file):
        book = load_workbook(output_file)
        if 'dim_audio' in book.sheetnames:
            book.remove(book['dim_audio'])
        book.save(output_file)
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
            audio_df.to_excel(writer, sheet_name='dim_audio', index=False)
    else:
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
            audio_df.to_excel(writer, sheet_name='dim_audio', index=False)
except PermissionError:
    print(f"PermissionError: Cannot write to '{output_file}'. Ensure the file is not open in another application (e.g., Excel) and you have write permissions.")
    exit(1)
except Exception as e:
    print(f"Error writing to Excel file '{output_file}': {e}")
    exit(1)
