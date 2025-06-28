import pandas as pd
import os
import json
import uuid
from openpyxl import load_workbook

input_file = r"C:\Users\user\Downloads\raw_data.xlsx"

output_file = r"C:\Users\user\Desktop\Docs\final_data1.xlsx"

output_dir = os.path.dirname(output_file)
if not os.path.exists(output_dir):
    try:
        os.makedirs(output_dir)
        print(f"Created directory: {output_dir}")
    except Exception as e:
        print(f"Error creating directory {output_dir}: {e}")
        exit(1)

try:
    df = pd.read_excel(input_file, sheet_name='Sheet1')
except FileNotFoundError:
    print(f"Error: Input file '{input_file}' not found.")
    exit(1)
except Exception as e:
    print(f"Error reading input file '{input_file}': {e}")
    exit(1)

comm_data = []

seen_source_ids = set()
for index, row in df.iterrows():
    try:
        content = json.loads(row['raw_content'])
        if 'meeting_attendees' in content:
            seen_emails = set()
            unique_attendees = []
            for attendee in content['meeting_attendees']:
                email = attendee.get('email')
                if email and email not in seen_emails:
                    seen_emails.add(email)
                    unique_attendees.append(attendee)
            content['meeting_attendees'] = unique_attendees

        source_id = row['source_id']
        if source_id not in seen_source_ids:
            seen_source_ids.add(source_id)
            comm_id = str(uuid.uuid4())
            raw_id = None
            comm_type_id = 1 if index % 2 == 0 else 3
            subject_id = index + 1
            calendar_id = index + 1
            audio_id = index + 1
            video_id = None
            transcript_id = index + 1
            datetime_id = content.get('dateString') or '2025-03-01T00:00:00'
            ingested_at = '2025-06-12T13:29:09'
            processed_at = '2025-03-22T11:37:50' if index % 3 == 0 else None
            is_processed = True
            raw_title = content.get('title', f'Communication {index + 1}')
            raw_duration = str(content.get('duration', '0,000000'))
            comm_data.append({
                'comm_id': comm_id,
                'raw_id': raw_id,
                'source_id': source_id,
                'comm_type_id': comm_type_id,
                'subject_id': subject_id,
                'calendar_id': calendar_id,
                'audio_id': audio_id,
                'video_id': video_id,
                'transcript_id': transcript_id,
                'datetime_id': datetime_id,
                'ingested_at': ingested_at,
                'processed_at': processed_at,
                'is_processed': is_processed,
                'raw_title': raw_title,
                'raw_duration': raw_duration
            })
    except json.JSONDecodeError as e:
        print(f"Warning: Could not parse JSON in row {index + 2} (source_id: {row['source_id']}): {row['raw_content'][:50]}... (Error: {e})")
        continue
    except Exception as e:
        print(f"Unexpected error in row {index + 2} (source_id: {row['source_id']}): {e}")
        continue

comm_df = pd.DataFrame(comm_data)

try:
    if os.path.exists(output_file):
        book = load_workbook(output_file)
        if 'fact_communication' in book.sheetnames:
            book.remove(book['fact_communication'])
        book.save(output_file)
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:
            comm_df.to_excel(writer, sheet_name='fact_communication', index=False)
    else:
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
            comm_df.to_excel(writer, sheet_name='fact_communication', index=False)
except PermissionError:
    print(f"PermissionError: Cannot write to '{output_file}'. Ensure the file is not open in another application (e.g., Excel) and you have write permissions.")
    exit(1)
except Exception as e:
    print(f"Error writing to Excel file '{output_file}': {e}")
    exit(1)
